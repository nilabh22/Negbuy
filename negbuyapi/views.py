<<<<<<< HEAD
from itertools import chain
import PIL.Image as Image
import string
import pybase64
import io
import threading
import re
import geopy.distance
from geopy.geocoders import Nominatim
import csv
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.db.models import Q
# from drf_api_logger import API_LOGGER_SIGNAL
from .models import *
import random
import requests
# import bs4sta
=======
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.db.models import Q
from drf_api_logger import API_LOGGER_SIGNAL
from .models import *
import random
import requests
import bs4
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
import datetime
import os
# import os.path
from .contactus import contact_function
from .serializers import *
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
<<<<<<< HEAD

=======
import openpyxl
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5

@api_view(['POST'])
def login(request):
    user_id = request.headers['User-id']
    phone = request.data['phone']

    try:
        usr = userDB.objects.get(user_id=user_id)
        response = {
            'status': False,
            'user_id': usr.user_id,
            'phone': usr.phone,
            'first_name': usr.first_name,
            'last_name': usr.last_name,
            'language': '',
            'user_bio': 0 if usr.first_name == None else 1,
        }

    except:
        userDB.objects.create(
            user_id=user_id,
            phone=phone
        )
        response = {
            'status': True,
            'user_id': user_id,
            'phone': phone,
            'first_name': '',
            'last_name': '',
        }

    finally:
        return Response(response, status=200)


@api_view(['POST'])
def seller_login(request):
    phone = request.data['phone']
    password = request.data['password']

    try:
        usr = userDB.objects.get(role='Seller', phone=phone, password=password)
        try:
            usrBank = bankDetail.objects.get(user=usr)
            accountName = usrBank.accountName
            accountNumber = usrBank.accountNumber
            accountIfsc = usrBank.accountIfsc
        except:
            accountName = None
            accountNumber = None
            accountIfsc = None
        response = {
            'status': True,
            'user_id': usr.user_id,
            'phone': usr.phone,
            'password': usr.password,
            'seller_name': usr.seller_name,
            'date_of_birth': usr.date_of_birth,
            'company': usr.company,
            'gst_number': usr.gst_number,
            'account_name': accountName,
            'account_number': accountNumber,
            'account_ifsc': accountIfsc
        }

    except Exception as e:
        response = {
            'status': False,
            'message': 'Incorrect phone or password',
            'error_msg': str(e)
        }

    finally:
        return Response(response, status=200)


@api_view(['POST'])
def seller_signup(request):
    user_id = request.headers['User-id']
    password = request.data['password']
    phone = request.data['phone']

    try:
        usr = userDB.objects.get(Q(user_id=user_id) | Q(phone=phone))
        sameField = 'User Id ' if usr.user_id == user_id else 'Phone number '
        response = {
            'status': False,
            'message': sameField + 'already exists'
        }

    except:
        userDB.objects.create(
            user_id=user_id,
            password=password,
            phone=phone,
            role="Seller"
        )
        response = {
            'status': True,
            'user_id': user_id,
            'phone': phone,
            'password': password,
            'seller_name': '',
        }

    finally:
        return Response(response, status=200)


def addProduct(request, user):
<<<<<<< HEAD
    price_choice = request.data['price_choice']
=======
    price_choice = request.data['price_choice'].lower()
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5

    category = request.data['category']
    try:
        category_record = productCategory.objects.get(name=category)
    except:
        category_record = productCategory.objects.create(name=category)

<<<<<<< HEAD
    if price_choice == 'Add Price':
=======
    if price_choice == 'add price':
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
        product_record = product.objects.create(
            user=user,
            name=request.data['name'],
            category_id=category_record,
            brand=request.data['brand'],
<<<<<<< HEAD
            main_image=request.data['main_image'],
=======
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
            keyword=request.data['keywords'],
            color=request.data['colors'],
            size=request.data['size'],
            details=request.data['details'],
            price_choice=price_choice,
            price=request.data['price'],
            mrp=request.data['mrp'],
            sale_price=request.data['sale_price'],
            sale_startdate=request.data['sale_startdate'],
            sale_enddate=request.data['sale_enddate'],
            manufacturing_time=request.data['manufacturing_time'],
            maximum_order_quantity=request.data['maximum_order_quantity'],
<<<<<<< HEAD
            # terms = terms_record,
=======
            #terms = terms_record,
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
            weight=request.data['weight'],
            transportation_port=request.data['transportation_port'],
            packing_details=request.data['packing_details'],
            packing_address=request.data['packing_address']
<<<<<<< HEAD

=======
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
        )
        product_record.save()
        images = dict((request.data).lists())['image']
        for image in images:
            productImages.objects.create(product=product_record, image=image)

<<<<<<< HEAD
    elif price_choice == 'Price according to quantity':
=======
    elif price_choice == 'price according to quantity':
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
        product_record = product.objects.create(
            user=user,
            name=request.data['name'],
            category_id=category_record,
<<<<<<< HEAD
            brand=request.data['brand'],
            main_image=request.data['main_image'],
=======
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
            keyword=request.data['keywords'],
            color=request.data['colors'],
            size=request.data['size'],
            details=request.data['details'],
            price_choice=price_choice,
            quantity_price=request.data['quantity_price'],
<<<<<<< HEAD
            # terms = terms_record,
=======
            #terms = terms_record,
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
            weight=request.data['weight'],
            transportation_port=request.data['transportation_port'],
            packing_details=request.data['packing_details'],
            packing_address=request.data['packing_address']
<<<<<<< HEAD

=======
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
        )
        product_record.save()
        images = dict((request.data).lists())['image']
        for image in images:
            productImages.objects.create(product=product_record, image=image)


@api_view(['POST'])
def product_upload_api(request):
    user_id = request.headers['User-id']
    try:
        user = userDB.objects.get(user_id=user_id, role='Seller')
        addProduct(request, user)
        return Response({'success': 'Product Added'}, status=200)
    except Exception as e:
        return Response({'status': 'Error', 'error_msg': str(e)})


def getProductObject(product):
    try:
        inventory = product.inventory_id.quantity
    except:
        inventory = ''

    try:
        productCategory = product.category_id.name
    except:
        productCategory = ''

    try:
        prodImages = productImages.objects.filter(product=product)
        imageURL = []
        for prodImage in prodImages:
            imageURL.append(prodImage.image.url)
    except:
        imageURL = []

    detailsList = []
    list = [str.strip() for str in product.details.split(',')]
    for eachDetail in list:
        keyValue = eachDetail.split(':')
        detailsObject = {
            'title': keyValue[0].strip(),
            'description': keyValue[1].strip(),
        }
        detailsList.append(detailsObject)

    object = {
        'id': product.id,
        'name': product.name,
        'sku': product.sku,
        'category': product.category_id.name,
        'inventory': inventory,
        'main_image': str(product.main_image),
        'image': imageURL,
        'featured_products': product.featured_products,
        'best_selling_products': product.best_selling_products,
        'hot_selling_products': product.hot_selling_products,
        'fast_dispatch': product.fast_dispatch,
        'ready_to_ship': product.ready_to_ship,
        'customized_product': product.customized_product,
        'brand': product.brand,
        'keyword': [str.strip() for str in product.keyword.split(',')],
        'color': [str.strip() for str in product.color.split(',')],
        'size': [str.strip() for str in product.size.split(',')],
        'details': detailsList,
        'price_choice': product.price_choice,
        'price': product.price,
        'mrp': product.mrp,
        'sale_price': product.sale_price,
        'sale_startdate': product.sale_startdate,
        'sale_enddate': product.sale_enddate,
        'manufacturing_time': product.manufacturing_time,
        'quantity_price': product.quantity_price,
        'maximum_order_quantity': product.maximum_order_quantity,
        'weight': product.weight,
        'transportation_port': product.transportation_port,
        'packing_details': product.packing_details,
        'packing_address': product.packing_address,
        'status': product.status,
        'created_at': product.created_at,
        'modified_at': product.modified_at,
        'deleted_at': product.deleted_at,
        'rating': {
            'rate': 3.0,
            'count': 430
        }
    }
    return object


@api_view(['POST'])
def product_info(request):
    try:
<<<<<<< HEAD
        review_list = []
        product_id = request.data['product_id']
        product_info = product.objects.get(id=product_id)
        product_detail = product_detail_db.objects.filter(
            product__id=product_id)
        product_reviews = review_db.objects.filter(product__id=product_id)
        product_serialized = ProductSerializer(product_info).data
        # try:
        #     review_serialized = ReviewSerializer(
        #         product_reviews, many=True).data
        #     for i in review_serialized:
        #         review_list.append(i['rating'])

        #     review_dict = {
        #         'rating': round(sum(review_list)/len(review_list)),
        #         'count': len(review_list)
        #     }

        # except Exception as e:
        #     review_dict = {
        #         'rating': 0,
        #         'count': 0
        #     }

        # product_serialized['reviews'] = review_dict

        obj = {
            'status': 'True',
            'message': 'success',
            'data': product_serialized,
        }

        return Response(obj)
    except Exception as e:
        return Response({'error_msg': str(e)}, status=500)
=======
        product_id = request.data['product_id']
        product_info = product.objects.get(id=product_id)
        product_object = getProductObject(product_info)
        return Response(product_object, status=200)
    except Exception as e:
        return Response({'status': 'Product id does not exists', 'error_msg': str(e)})
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5


@api_view(['GET'])
def featured_product_api(request):
<<<<<<< HEAD
    featured_products = product.objects.filter(featured_products=True)
    featured_serialized = ProductSerializer(featured_products, many=True).data

    # reviews = review_db.objects.filter(
    #     id__in=featured_products.values_list('id', flat=True))
    # reviews_serialized = ReviewSerializer(reviews, many=True).data

    return Response({
        'Message': 'OK',
        'data': featured_serialized
    })
=======
    try:
        featured_product_list = []
        featured_products = product.objects.filter(featured_products=True)

        if featured_products.exists():
            for each_product in featured_products:
                featured_product_list.append(getProductObject(each_product))
            # if len(featured_product_list) != 10:
            #     products = list(product.objects.filter(
            #         featured_products=False))
            #     random_products = random.sample(
            #         products, 10-len(featured_product_list))
            #     for each_product in random_products:
            #         featured_product_list.append(
            #             getProductObject(each_product))

            return Response({
                'status': 'True',
                'message': 'Success',
                'data': featured_product_list
            })

        else:
            return Response({
                'status': 'True',
                'message': 'Success',
                'data': []
            })

    except Exception as e:
        return Response({
            'status': 'Error',
            'message': e,
            'data': ''
        })
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5


@api_view(['GET'])
def fast_dispatch_api(request):
<<<<<<< HEAD
    fast_dispatch_products = product.objects.filter(fast_dispatch=True)

    fast_dispatch_serialized = ProductSerializer(
        fast_dispatch_products, many=True).data

    return Response({
        'status': 200,
        'message': 'success',
        'data': fast_dispatch_serialized
    })
=======
    fast_dispatch_list = []
    fast_dispatch_products = product.objects.filter(fast_dispatch=True)[:10]
    for each_product in fast_dispatch_products:
        fast_dispatch_list.append(getProductObject(each_product))

    if len(fast_dispatch_list) != 10:
        products = list(product.objects.filter(fast_dispatch=False))
        random_products = random.sample(
            products, 10-len(fast_dispatch_list))
        for each_product in random_products:
            fast_dispatch_list.append(getProductObject(each_product))
    return Response(fast_dispatch_list, status=200)
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5


@api_view(['GET'])
def ready_to_ship_api(request):
<<<<<<< HEAD
    ready_to_ship_products = product.objects.filter(ready_to_ship=True)

    ready_to_ship_products_serialized = ProductSerializer(
        ready_to_ship_products, many=True).data

    return Response({
        'status': 200,
        'message': 'success',
        'data': ready_to_ship_products_serialized
    })
=======
    ready_to_ship_list = []
    ready_to_ship_products = product.objects.filter(ready_to_ship=True)[:10]
    for each_product in ready_to_ship_products:
        ready_to_ship_list.append(getProductObject(each_product))

    if len(ready_to_ship_list) != 10:
        products = list(product.objects.filter(ready_to_ship=False))
        random_products = random.sample(
            products, 10-len(ready_to_ship_list))
        for each_product in random_products:
            ready_to_ship_list.append(getProductObject(each_product))
    return Response(ready_to_ship_list, status=200)
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5


@api_view(['GET'])
def customized_product_api(request):
<<<<<<< HEAD
    customized_products = product.objects.filter(customized_product=True)
    customized_products_serialized = ProductSerializer(
        customized_products, many=True).data

    return Response({
        'status': 200,
        'message': 'success',
        'data': customized_products_serialized
    })
=======
    customized_product_list = []
    customized_products = product.objects.filter(customized_product=True)[:10]
    for each_product in customized_products:
        customized_product_list.append(getProductObject(each_product))

    if len(customized_product_list) != 10:
        products = list(product.objects.filter(customized_product=False))
        random_products = random.sample(
            products, 10-len(customized_product_list))
        for each_product in random_products:
            customized_product_list.append(getProductObject(each_product))
    return Response(customized_product_list, status=200)
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5


@api_view(['GET'])
def new_arrivals_api(request):
<<<<<<< HEAD
    new_arrivals_products = product.objects.all().order_by('-id')[:10]

    newArrival_products_serialized = ProductSerializer(
        new_arrivals_products, many=True).data
    return Response({
        'status': 200,
        'message': 'success',
        'data': newArrival_products_serialized
    })
=======
    new_arrivals_list = []
    new_arrivals_products = product.objects.all().order_by('-id')[:10]
    for each_product in new_arrivals_products:
        new_arrivals_list.append(getProductObject(each_product))
    return Response(new_arrivals_list, status=200)
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5


@api_view(['GET'])
def top_selling_api(request):
<<<<<<< HEAD
    top_selling_products = list(product.objects.all())
    random_top_selling_products = random.sample(top_selling_products, 10)

    topSelling_products_serialized = ProductSerializer(
        random_top_selling_products, many=True).data
    return Response({
        'status': 200,
        'message': 'success',
        'data': topSelling_products_serialized
    })
=======
    top_selling_list = []
    top_selling_products = list(product.objects.all())
    random_top_selling_products = random.sample(top_selling_products, 10)
    for each_product in random_top_selling_products:
        top_selling_list.append(getProductObject(each_product))
    return Response(top_selling_list, status=200)
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5


@api_view(['POST'])
def add_to_cart(request):
    user_id = request.headers['User-id']

    try:
        usr = userDB.objects.get(user_id=user_id)
        product_id = int(request.data['product_id'])
        quantity = int(request.data['quantity'])

        try:
            record = cart.objects.get(user_id=usr.id, product_id=product_id)
            record.quantity = quantity
            record.save()
        except:
            cart.objects.create(
                user_id=usr.id,
                product_id=product_id,
                quantity=quantity,
            )
        return Response({'status': 'success'}, status=200)

    except Exception as e:
        return Response({'status': 'error', 'error_msg': str(e)}, status=401)


@api_view(['POST'])
def remove_from_cart(request):
    user_id = request.headers['User-id']

    try:
        usr = userDB.objects.get(user_id=user_id)
        product_id = int(request.data['product_id'])
        removed_quantity = int(request.data['removed_quantity'])
        item_quantity = cart.objects.get(
            user_id=usr.id, product_id=product_id).quantity
        new_quantity = item_quantity - removed_quantity

        if new_quantity <= 0:
            cart.objects.get(user_id=usr.id, product_id=product_id).delete()
            return Response({'status': 'success', 'msg': 'item removed from cart'})
        else:
            record = cart.objects.get(user_id=usr.id, product_id=product_id)
            record.quantity = new_quantity
            record.save()
            return Response({'status': 'success', 'msg': 'item quantity decreased to ' + str(new_quantity)})

    except Exception as e:
        return Response({'status': 'error', 'error_msg': str(e)})


@api_view(['GET'])
def my_cart(request):
    status = 200
    cart_items = []
    user_id = request.headers['User-id']
    try:
        usr = userDB.objects.get(user_id=user_id)
        all_items = cart.objects.filter(user_id=usr.id)
        for each_item in all_items:
            object = getProductObject(each_item.product)
            object['quantity'] = each_item.quantity
            cart_items.append(object)
    except:
        status = 401

    return Response(cart_items, status=status)


@api_view(['GET'])
def user_products_api(request):
    user_id = request.headers['User-id']
    response = []
<<<<<<< HEAD
    # try:
    user = userDB.objects.get(user_id=user_id, role='Seller')
    if user:
        user_products = product.objects.filter(user=user).order_by('-id')
        user_products_serialized = ProductSerializer(user_products, many=True)
        return Response({
            'status': True,
            'message': 'Success',
            'data': user_products_serialized.data
        })
    else:
        return Response(response, status=200)
    # except Exception as e:
    #     return Response({'status': 'error', 'error_msg': str(e)}, status=401)
=======
    try:
        user = userDB.objects.get(user_id=user_id, role='Seller')
        user_products = product.objects.filter(user=user)
        for each_product in user_products:
            response.append(getProductObject(each_product))
    except Exception as e:
        return Response({'status': 'error', 'error_msg': str(e)}, status=401)

    return Response(response, status=200)
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5


def getGstObject(gst_number, divData):
    object = {
        'gst_number': gst_number,
        'tradeName': divData[0].input.get('value'),
        'legalName': divData[1].input.get('value'),
        'type': divData[2].input.get('value'),
        'regDate': divData[3].input.get('value'),
        'constBusiness': divData[4].input.get('value'),
        'businessNature': divData[5].input.get('value'),
        'principalPlace': divData[6].input.get('value'),
    }
    return object


@api_view(['POST'])
def verify_gst(request):
    user_id = request.headers['User-id']
    gst_number = request.data['gstNo']

    try:
<<<<<<< HEAD
        usr = userDB.objects.get(user_id=user_id, role='Seller')
=======
        usr = userDB.objects.get(user_id=user_id)
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
        addr = "https://irisgst.com/gstin-filing-detail/?gstinno=" + gst_number
        response = requests.get(addr)
        htmlPage = bs4.BeautifulSoup(response.text, "html.parser")
        divData = htmlPage.find_all('div', {'class': 'form-group'})
        response = getGstObject(gst_number, divData)
        usr.gst_number = gst_number
        usr.save(update_fields=['gst_number'])
        return Response(response, status=200)

    except:
        return Response({'status': 'Invalid GST Number'}, status=403)


@api_view(['POST'])
def bank_details(request):
    user_id = request.headers['User-id']

    try:
        usr = userDB.objects.get(user_id=user_id, role='Seller')
        accName = request.data['accName']
        accNo = request.data['accNo']
        ifsc = request.data['ifsc']

        bankDetail.objects.create(
            user_id=usr.id,
            accountName=accName,
            accountNumber=accNo,
            accountIfsc=ifsc
        )
        return Response({'status': 'success'}, status=200)

    except:
        return Response({'status': 'User does not exist'}, status=401)


@api_view(['POST'])
def seller_details(request):
    user_id = request.headers['User-id']

    try:
        usr = userDB.objects.get(user_id=user_id, role='Seller')
        usr.seller_name = request.data['seller_name']
        usr.date_of_birth = request.data['date_of_birth']
        usr.email = request.data['email']
        usr.company = request.data['company']
        usr.address = request.data['address']
        usr.document_verification = request.FILES['document_verification']
        usr.save(update_fields=['seller_name', 'date_of_birth',
                 'email', 'company', 'address', 'document_verification'])
        return Response({'status': 'success'}, status=200)

    except Exception as e:
        return Response({'status': 'error', 'error_msg': str(e)}, status=401)


@api_view(['POST'])
def search_category(request):
    response = []

    category_selected = request.data['category_selected']
    raw_string = request.data['category']
    file_path = 'static/categories/'+str(category_selected).strip()+'.txt'

    if (not (raw_string and raw_string.strip())):
        keywords = raw_string.replace('>', '').replace(
            '& ', '').replace('and ', '').strip()

<<<<<<< HEAD
=======
    print(category_selected, raw_string, keywords)

>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
    if os.path.exists(file_path):
        if len(keywords) == 0:
            with open(file_path) as file:
                catLines = file.readlines()
                for line in catLines:
                    response.append(line.strip())
            return Response({
                'status': True,
                'message': 'Success',
                'data': response
            })

        else:
            file_path = 'static/categories/' + \
                str(category_selected).strip()+'.txt'
<<<<<<< HEAD
=======
            print(file_path)
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
            with open(file_path) as file:
                for line in file:
                    if keywords.lower() in line.lower().replace('& ', ''):
                        response.append(line.strip())
                if len(response) == 0:
                    keywords = keywords.replace(',', '')
                    keywordList = keywords.split(' ')
                    with open(file_path) as file:
                        for line in file:
                            if any(keyword.lower() in line.lower() for keyword in keywordList):
                                response.append(line.strip())
            return Response(response, status=200)
    else:
        return Response({
            'status': True,
            'message': 'Success',
            'data': 'Invalid Category'
        })


<<<<<<< HEAD
@api_view(['GET'])
def get_ports(request):
    all_ports = port.objects.filter(country='Brazil')
    port_serializer = PortSerializer(all_ports, many=True)
    return Response({
        'status': True,
        'message': 'Success',
        'data': port_serializer.data
    })
=======
def getPort(port):
    name = port.name
    state = port.state
    return name + ',' + state


@api_view(['GET'])
def get_ports(request):
    response = []
    all_ports = port.objects.all()
    for each_port in all_ports:
        response.append(getPort(each_port))
    return Response(response, status=200)
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5


@api_view(['GET'])
def get_categories(request):
    try:
        # /home/vf2586e813kg/api.negbuy/static
<<<<<<< HEAD
        path = "/home/negbswof/api.negbuy/static/categories"
=======
        path = "/home/vf2586e813kg/api.negbuy/static/categories"
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
        # path = "/Users/habibahmed/Desktop/NegBuy/negbuy/static/categories"
        dir_list = os.listdir(path)
        mylst = map(lambda each: each.strip(".txt"), dir_list)

<<<<<<< HEAD
        # for item in mylst:
        #     productCategory.objects.create(
        #         name=item,
        #         desc="Lorem Ipsum"
        #     )

=======
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
        return Response({
            'status': True,
            'message': 'Success',
            'data': mylst
        })
    except Exception as e:
        return Response({
            'status': False,
            'message': e,
            'data': ''
        })


@api_view(['GET'])
def get_orders(request):
    user_id = request.headers['User-id']
    all_orders_list = list()
    try:
        usr = userDB.objects.get(user_id=user_id, role='Seller')
        # get orders details
        all_orders = orders.objects.filter(user=usr)
        for each_order in all_orders:
            obj = {
                'order_date': each_order.created_at,
                'order_number': each_order.order_number,
                'product_name': each_order.product_info.name,
                'product_quantity': each_order.order_quantity,
                'ship_date': each_order.shipping_date,
                'delivery_date': each_order.delivery_date,
                'order_status': each_order.status
            }

            all_orders_list.append(obj)
        return Response({
            'status': True,
            'message': 'Success',
            'data': all_orders_list
        })
    except Exception as e:
        return Response({
            'status': False,
            'message': e,
            'data': ''
        })


@api_view(['POST'])
def contactus_function(request):
    try:
        name = request.data['name']
        contact_num = request.data['number']
        req_email = request.data['email']
        message = request.data['message']
        full_message = "Please contact "+name+" " + \
            contact_num+" "+req_email+" for "+message
        cont_res = contact_data.objects.create(message=full_message)
        if cont_res:
            return Response({
                'status': True,
                'message': 'Success',
                'data': "Message sent"
            })
        else:
            return Response({
                'status': False,
                'message': 'Error',
                'data': "Please enter details again"
            })
    except:
        return Response({
            'status': False,
            'message': 'Error',
            'data': "Please enter details again"
        })


@api_view(['POST'])
def delete_product(request):
    try:
        user = request.headers['User-id']
        product_id = request.data['product_id']

        user_details = userDB.objects.get(user_id=user)
        product_details = product.objects.get(
            id=product_id, user=user_details).delete()

        return Response({
            'status': True,
            'message': 'Success',
            'data': "Deleted successfully"
        })
    except Exception as e:
        return Response({
            'status': False,
            'message': 'Error',
            'data': str(e)
        })


@api_view(['POST'])
def product_detail(request):
    try:
        product_id = request.data['product_id']
<<<<<<< HEAD
=======
        response = []
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
        image_list = []

        product_details = product.objects.filter(id=product_id)
        image_details = productImages.objects.filter(product__id=product_id)
<<<<<<< HEAD
        product_db_details = product_detail_db.objects.filter(
            product__id=product_id)

        serializer = ProductSerializer(product_details, many=True)
        productDetails_serializer = ProductDetailSerializer(
            product_db_details, many=True)
=======

        serializer = ProductSerializer(product_details, many=True)
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
        img_serializer = ImageSerializer(image_details, many=True)
        for dic in img_serializer.data:
            for key in dic:
                if key == 'image':
                    image_list.append(dic[key])

<<<<<<< HEAD
        data_dict = {
            'product details': serializer.data,
            'image detail': productDetails_serializer.data,
            'images': image_list
        }

        return Response({
            'status': True,
            'message': 'Success',
            'data': data_dict
=======
        return Response({
            'status': True,
            'message': 'Success',
            'data': serializer.data,
            'images': image_list
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
        })

    except Exception as e:
        return Response({
            'status': 'False',
            'message': 'Error',
            'data': str(e)
        })


<<<<<<< HEAD
@ api_view(['GET'])
def best_selling(request):
    try:
        best_selling = product.objects.filter(best_selling_products=True)[:10]
        bestSelling_serialized = ProductSerializer(
            best_selling, many=True).data

        return Response({
            'status': 'True',
            'message': 'Success',
            'data': bestSelling_serialized
        })
=======
@api_view(['GET'])
def best_selling(request):
    try:
        best_selling_list = []
        best_selling = product.objects.filter(best_selling_products=True)[:10]
        if best_selling.exists():
            for each_product in best_selling:
                best_selling_list.append(getProductObject(each_product))

            if len(best_selling_list) != 10:
                products = list(product.objects.filter(
                    best_selling_products=False))
                random_products = random.sample(
                    products, 10-len(best_selling_list))
                for each_product in random_products:
                    best_selling_list.append(getProductObject(each_product))

            return Response({
                'status': 'True',
                'message': 'Success',
                'data': best_selling_list
            })

        else:
            return Response({
                'status': 'True',
                'message': 'Success',
                'data': []
            })
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5

    except Exception as e:
        return Response({
            'status': 'Error',
            'message': e,
            'data': ''
        })


<<<<<<< HEAD
@ api_view(['GET'])
def hot_selling(request):
    try:
        hot_selling = product.objects.filter(hot_selling_products=True)[:10]
        hotSelling_serialized = ProductSerializer(hot_selling, many=True).data
        return Response({
            'status': 'True',
            'message': 'Success',
            'data': hotSelling_serialized
        })
=======
@api_view(['GET'])
def hot_selling(request):
    try:
        hot_selling_list = []
        hot_selling = product.objects.filter(hot_selling_products=True)[:10]

        if hot_selling.exists():
            for each_product in hot_selling:
                hot_selling_list.append(getProductObject(each_product))

            # Adding random Products when out of hot_selling_products
            if len(hot_selling_list) != 10:
                products = list(product.objects.filter(
                    hot_selling_products=False))
                random_products = random.sample(
                    products, 10-len(hot_selling_list))
                for each_product in random_products:
                    hot_selling_list.append(getProductObject(each_product))

            return Response({
                'status': 'True',
                'message': 'Success',
                'data': hot_selling_list
            })
        else:
            return Response({
                'status': 'True',
                'message': 'Success',
                'data': []
            })
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5

    except Exception as e:
        return Response({
            'status': 'Error',
            'message': e,
            'data': ''
        })


<<<<<<< HEAD
@ api_view(['POST'])
def categorized_product(request):
    try:
        category = request.data['category']
=======
@api_view(['POST'])
def categorized_product(request):
    try:
        category = request.data['category']
        imageList = []
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5

        product_list = product.objects.filter(category_id__name=category)
        serializer = ProductSerializer(product_list, many=True)

        return Response({
            'status': True,
            'message': 'Success',
            'data': serializer.data,
        })

    except Exception as e:
        return Response({
            'status': 'Error',
            'message': e,
            'data': ''
        })
# =================================================================================#

<<<<<<< HEAD

# new api of read file product_list.json.....


@ api_view(['GET'])
def read_json(request):
    # remove on live version
    # all_products = productCategory.objects.all().delete()
    # all_inventory = productInventory.objects.all().delete()
    user_data_all = userDB.objects.all()
    category_data = productCategory.objects.all()

    bool_list = [True, False]
    with open('product_lists.json', 'r') as f:
        jsondata = f.read()
        obj = json.loads(jsondata)
        user_id = ['Neyyjn9DCncSDQDWtJNEa57u6el2',
                   'M1TOJ78YaQcGJZ6fsuMoxCB0ad93', '2utKoA5gheWtqe5ZRBNqdFamtQr1']
        # user_id = ['Neyyjn9DCncSDQDWtJNEa57u6el2']
        for pd in obj:

            # inventory_data = Inventory.objects.create(quantity=20)
            # inventory_data.save()

            user_data = random.choice(user_data_all)
            category_obj = random.choice(category_data)
            # inventory_obj = inventory_data
            name = str(pd['name'])
            sku = str(pd['sku'])
            main_image = str(pd['main_image'])
            hot_selling_products = random.choice(bool_list)
            best_selling_products = random.choice(bool_list)
            featured_products = random.choice(bool_list)
            fast_dispatch = random.choice(bool_list)
            ready_to_ship = random.choice(bool_list)
            customized_product = random.choice(bool_list)
=======
# new api of read file product_list.json.....
import random

@api_view(['GET'])
def read_json(request):
    #remove on live version
    #all_products = productCategory.objects.all().delete()
    #all_inventory = productInventory.objects.all().delete()
    user_data_all = userDB.objects.all()
    category_data = productCategory.objects.all()

    with open('EG_products.json', 'r') as f:
        jsondata = f.read()
        obj = json.loads(jsondata)
        for pd in obj:
            
            inventory_data = productInventory.objects.create(quantity=20)
            inventory_data.save()

            user_data = random.choice(user_data_all)
            category_obj = random.choice(category_data)
            inventory_obj = inventory_data
            name = str(pd['name'])
            sku = str(pd['sku'])
            featured_products = str(pd['featured_products'])
            fast_dispatch = str(pd['fast_dispatch'])
            ready_to_ship = str(pd['ready_to_ship'])
            customized_product = str(pd['customized_product'])
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
            brand = str(pd['brand'])
            keyword = str(pd['keyword'])
            color = str(pd['color'])
            size = str(pd['size'])
            details = str(pd['details'])
            price_choice = str(pd['price_choice'])
            price = str(pd['price'])
            mrp = str(pd['mrp'])
            sale_price = str(pd['sale_price'])
            sale_startdate = str(pd['sale_startdate'])
            sale_enddate = str(pd['sale_enddate'])
            manufacturing_time = str(pd['manufacturing_time'])
            quantity_price = str(pd['quantity_price'])
            maximum_order_quantity = str(pd['maximum_order_quantity'])
            weight = str(pd['weight'])
            transportation_port = str(pd['transportation_port'])
            packing_details = str(pd['packing_details'])
            packing_address = str(pd['packing_address'])
            status = str(pd['status'])
            created_at = str(pd['created_at'])
            modified_at = str(pd['modified_at'])
            deleted_at = str(pd['deleted_at'])

<<<<<<< HEAD
            product_obj = product.objects.create(
                user=user_data,
                name=name,
                sku=sku,
                main_image=main_image,
                category_id=category_obj,
                # inventory_id=inventory_obj,
                hot_selling_products=hot_selling_products,
                best_selling_products=best_selling_products,
                featured_products=True,
=======
            product.objects.create(
                user=user_data,
                name=name,
                sku=sku,
                category_id=category_obj,
                inventory_id=inventory_obj,
                featured_products=featured_products,
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
                fast_dispatch=fast_dispatch,
                ready_to_ship=ready_to_ship,
                customized_product=customized_product,
                brand=brand,
                keyword=keyword,
                color=color,
                size=size,
                details=details,
                price_choice=price_choice,
<<<<<<< HEAD
                price=0 if price == 'None' else float(price),
                mrp=0 if mrp == 'None' else float(mrp),
=======
                price=price,
                mrp=mrp,
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
                sale_price=sale_price,
                sale_startdate=sale_startdate,
                sale_enddate=sale_enddate,
                manufacturing_time=manufacturing_time,
                quantity_price=quantity_price,
                maximum_order_quantity=maximum_order_quantity,
                weight=weight,
                transportation_port=transportation_port,
                packing_details=packing_details,
                packing_address=packing_address,
                status=status,
                created_at=created_at,
                modified_at=modified_at,
                deleted_at=deleted_at
            )
<<<<<<< HEAD
            product_obj.save()
            Inventory.objects.create(product=product_obj, quantity=0)
=======
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5

    return Response({
        'status': 'success',
        'message': 'Added products',
        'data': obj
    })


# ---------------------------- Seller.Negbuy -------------------------------------- #

<<<<<<< HEAD
# =========== DMS to DEC ============ #

def dms2dec(dms_str):

    dms_str = re.sub(r'\s', '', dms_str)

    sign = -1 if re.search('[swSW]', dms_str) else 1

    numbers = [*filter(len, re.split('\D+', dms_str, maxsplit=4))]

    degree = numbers[0]
    minute = numbers[1] if len(numbers) >= 2 else '0'
    second = numbers[2] if len(numbers) >= 3 else '0'
    frac_seconds = numbers[3] if len(numbers) >= 4 else '0'

    second += "." + frac_seconds
    return sign * (int(degree) + float(minute) / 60 + float(second) / 3600)


# Funtion to Read and add Port Details from XLSX
@ api_view(['POST'])
def add_ports(request):
    start_line = int(request.data['start_line'])
    end_line = int(request.data['end_line'])
    wb = load_workbook('portdata.xlsx')
    ws = wb.active

    for row in range(start_line, end_line+1):
        for col in range(1, 5):
            char = get_column_letter(col)
=======
# Function to Read and add Port Details from XLSX
@api_view(['POST'])
def add_ports(request):
    line_no = int(request.data['line_no'])
    wb = load_workbook('Ports Data.xlsx')
    ws = wb.active

    for row in range(2, line_no+1):
        for col in range(1, 5):
            char = get_column_letter(col)
            print(ws[char+str(row)].value)

>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
            country = ws[char+str(row)].value
            portname = ws[get_column_letter(col+1)+str(row)].value
            lat = ws[get_column_letter(col+2)+str(row)].value
            lon = ws[get_column_letter(col+3)+str(row)].value
<<<<<<< HEAD
            try:
                dec_lat = dms2dec(lat)
                dec_lon = dms2dec(lon)
            except Exception as e:
                print(e)
            port.objects.create(
                name=portname,
                country=country,
                latitude=dec_lat,
                longitude=dec_lon
=======

            port.objects.create(
                name=portname,
                country=country,
                latitude=lat,
                longitude=lon
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
            )
            break

    return Response({
        'status': 'success',
        'message': 'Added ports',
    })


# Api to fetch order with status Running
<<<<<<< HEAD
@ api_view(['POST'])
=======
@api_view(['POST'])
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
def my_orders(request):
    user_id = request.headers['User-id']
    running_orders = orders.objects.filter(status='Running', user__id=user_id)

    order_serializer = OrderSerializer(running_orders, many=True)

    return Response({
        'status': 'success',
        'message': 'Added ports',
        'data': order_serializer.data
    })


# Api to fetch orders with status Completed
<<<<<<< HEAD
@ api_view(['POST'])
=======
@api_view(['POST'])
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
def order_history(request):
    user_id = request.headers['User-id']
    completed_orders = orders.objects.filter(
        status='Completed', user__id=user_id)

    order_serializer = OrderSerializer(completed_orders, many=True)

    return Response({
        'status': 'success',
        'message': 'Added ports',
        'data': order_serializer.data
    })

<<<<<<< HEAD
# Add order Note to a particular order


@ api_view(['POST'])
def order_note(request):

    user_id = request.headers['User-id']
    order_id = request.data['order_id']
    note = request.data['note']

    order_obj = orders.objects.get(order_id=order_id)
    order_obj.order_note = note
    order_obj.save()

    return Response({
        'status': 'success',
        'message': 'Added ports',
        'data': "Success"
    })

# Get order Details


@ api_view(['GET'])
def order_details(request):
    user_id = request.headers['User-id']
    order_id = request.data['order_id']

    order = orders.objects.filter(id=order_id).values()
    order_item = orders.objects.get(id=order_id)
    product_detail = product.objects.filter(
        name=order_item.product_info).values()
    user_detail = userDB.objects.filter(user_id=user_id)

    order_serialized = OrderSerializer(order, many=True)
    product_serialized = ProductSerializer(product_detail, many=True)
    user_serialized = UserSerializer(user_detail, many=True)

    quantity = 0
    price = 0
    # Price calc
    for item in order:
        quantity = int(item.get('order_quantity'))

    for item in product_detail:
        price = item.get('price')

    cart_total = quantity * price

    data_dict = {
        'Product Details': product_serialized.data,
        'Order Details': order_serialized.data,
        'Customer Details': user_serialized.data,
        'Price Details': cart_total
    }

    return Response({
        'status': 'success',
        'message': 'Order Details',
        'data': data_dict
    })


# Image uploading of Products from CSV file
@ api_view(['GET'])
=======
import csv


@api_view(['GET'])
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
def read_csv(request):
    with open('images.csv', 'r') as csv_file:
        csv_reader = csv.reader(csv_file)
        for a in csv_reader:
            if product.objects.filter(name=a[0]).exists():
                product_object = product.objects.get(name=a[0])
                productImages.objects.create(
                    product=product_object,
                    image=a[1]
                )
    return Response({
        'status': 'success'
    })


<<<<<<< HEAD
# Distance between Longitude and latitude
@ api_view(['GET'])
def port_distance(request):
    a_lat = float(request.data['a_lat'])
    a_lon = float(request.data['a_lon'])

    geolocator = Nominatim(user_agent="geoapiExercises")

    location = geolocator.reverse(str(a_lat)+","+str(a_lon))
    address = location.raw['address']
    country = address.get('country').lower()
    dist = 0
    cl_port = ''
    coords_1 = (a_lat, a_lon)
    qs = port.objects.filter(country__icontains=country).values()
    for item in qs:
        latitude_val = item.get('latitude')
        longitude_val = item.get('longitude')
        coords_2 = (latitude_val, longitude_val)
        distance = geopy.distance.geodesic(coords_1, coords_2).km
        if dist < distance:
            dist = distance
            cl_port = item.get('name')
            data_dict = {
                'Distance': dist,
                'port': cl_port
            }
        return Response({
            'status': 'success',
            'message': 'Order Details',
            'data': data_dict
        })
    return Response({
        'status': 'success',
        'message': 'Order Details',
        'data': ''
    })


# Add reviews
@ api_view(['POST'])
def post_review(request):
    user_id = request.headers['user-id']
    review = request.data['review']
    review_title = request.data['review_title']
    product_id = request.data['product_id']
    rating = request.data['rating']
    try:
        user = userDB.objects.get(user_id=user_id)
        products = product.objects.get(id=product_id)

        review_db.objects.create(
            user=user,
            product=products,
            review=review,
            rating=rating,
            review_title=review_title
        )

        user_obj = userDB.objects.get(user_id=user_id)
        userSerialized = UserSerializer(user_obj)

        data_dict = {
            'user': userSerialized.data,
            'product': products.name,
            'review title': review_title,
            'review': review,
            'rating': int(rating)
        }

        return Response({
            'status': 'success',
            'message': 'Review posted',
            'data': data_dict
        })
    except Exception as e:
        return Response({
            'status': 'success',
            'message': str(e)
        })


# Get a list of reviews
@ api_view(['POST'])
def product_reviews(request):
    product_id = request.data['product_id']

    product_reviews = review_db.objects.filter(product__id=product_id)

    reviews_serialized = ReviewSerializer(product_reviews, many=True)

    return Response({
        'status': 'success',
        'message': 'Review posted',
        'data': reviews_serialized.data
    })


def saveImage(img):
    img.save('media/Product_images/' + img)
    return 'Product_images/' + str(img)


@ api_view(['POST'])
def upload_product(request):
    user = request.headers['User-id']
    main_images = request.FILES.getlist(
        'main_image')  # main_image ['', '', '']
    image_list = request.FILES.getlist('image')  # image ['','','']
    jdata = dict((request.data).lists())['data']
    data = json.loads(jdata[0])

    try:
        if data['screen_no'] == 1:
            if data['varients'] == []:
                user_obj = userDB.objects.get(user_id=user)
                category_obj = productCategory.objects.get(
                    name=data['category'])
                product_obj = product.objects.create(
                    user=user_obj,
                    category_id=category_obj,
                    sub_category=data['sub_category'],
                    name=data['name'],
                    brand=data['brand'],
                    keyword=str(data['keyword']).translate(
                        {ord(c): None for c in "'[]"}),
                    details=str(data['details']).translate(
                        {ord(c): None for c in "'[]"}),
                    size=str(data['size']).translate(
                        {ord(c): None for c in "'[]"}),
                    color=str(data['color']).translate(
                        {ord(c): None for c in "'[]"}),
                    main_image=main_images[0],
                    page_status=data['screen_no'],
                )
                product_obj.save()
                for img in image_list:
                    productImages.objects.create(
                        product=product_obj, image=img)
                return Response({
                    'status': 'success',
                    'message': 'product added',
                    'data': product_obj.id
                })

            else:
                user_obj = userDB.objects.get(user_id=user)
                category_obj = productCategory.objects.get(
                    name=data['category'])
                product_obj = product.objects.create(
                    user=user_obj,
                    category_id=category_obj,
                    sub_category=data['sub_category'],
                    name=data['name'],
                    brand=data['brand'],
                    keyword=str(data['keyword']).translate(
                        {ord(c): None for c in "'[]"}),
                    details=str(data['details']).translate(
                        {ord(c): None for c in "'[]"}),
                    varients=True,
                    page_status=data['screen_no'],
                )
                product_obj.save()
                for i in data['varients']:
                    color = i['color']
                    size = str(i['size']).translate(
                        {ord(c): None for c in "'[]"})

                    for m_img in main_images:
                        if str(m_img) in i['main_image']:
                            productImages.objects.create(
                                product=product_obj, image=m_img, main_image=True, size=size, color=color)
                            break

                    for img in image_list:
                        if str(img) in i['images']:
                            productImages.objects.create(
                                product=product_obj, image=img, main_image=False, size=size, color=color)

                return Response({
                    'status': 'success',
                    'message': 'product added',
                    'data': product_obj.id
                })

        elif data['screen_no'] == 2:
            product_obj = product.objects.get(id=data['product_id'])
            product_obj.detailed_description = data['detailed_description']
            product_obj.page_status = data['screen_no']
            product_obj.save()
            for i in data['product_details']:
                count = 0
                product_detail_db.objects.create(
                    product=product_obj,
                    image=image_list[count],
                    heading=i['title'],
                    desc=i['description']
                )
                count += 1

            return Response({
                'status': 'success',
                'message': 'product added',
                'data': product_obj.id
            })

        elif data['screen_no'] == 3:
            product_obj = product.objects.get(id=data['product_id'])

            if data['price_choice'] == 'Add Price':
                product_obj.price_choice = str(data['price_choice'])
                product_obj.price = data['price']
                product_obj.mrp = data['mrp']
                product_obj.sale_price = data['sale_price']
                product_obj.sale_startdate = data['sale_startdate']
                product_obj.sale_enddate = data['sale_enddate']
                product_obj.maximum_order_quantity = data['minimum_order_quantity']
                product_obj.manufacturing_time = data['manufacturing_time']
                product_obj.page_status = data['screen_no']
                product_obj.save()

                return Response({
                    'status': 'success',
                    'message': 'product added',
                    'data': product_obj.id
                })

            elif data['price_choice'] == 'Price according to quantity':
                product_obj.price_choice = str(data['price_choice'])
                product_obj.quantity_price = str(data['quantity_price']).translate(
                    {ord(c): None for c in "'[]"})
                product_obj.page_status = data['screen_no']
                product_obj.save()

                return Response({
                    'status': 'success',
                    'message': 'product added',
                    'data': product_obj.id
                })

        elif data['screen_no'] == 4:
            product_obj = product.objects.get(id=data['product_id'])
            product_obj.weight = data['weight']
            product_obj.transportation_port = data['transportation_port']
            product_obj.packing_details = data['packing_details']
            product_obj.packing_address = data['package_address']
            product_obj.page_status = data['screen_no']
            product_obj.save()
            if product_obj.varients:
                image_detail_obj = list(productImages.objects.filter(
                    product=product_obj).values_list('color', 'size'))
                colors = []
                for item in set(image_detail_obj):
                    inventory_obj = Inventory.objects.create(
                        product=product_obj,
                        quantity=0,
                        color=item[0],
                        size=item[1],
                    )
            else:
                inventory_obj = Inventory.objects.create(
                    product=product_obj,
                    quantity=0,
                    color=product_obj.color,
                    size=product_obj.size,
                )
                inventory_obj.save()

            return Response({
                'status': 'success',
                'message': 'product added',
                'data': 'Product uploaded'
            })
    except Exception as e:
        return Response({
            'status': 'error',
            'message': e,
        })


@ api_view(['POST'])
def video_upload(request):
    product_id = request.data['product_id']
    video = request.FILES['video']
    try:
        product_obj = product.objects.get(id=product_id)
        product_obj.video = video
        product_obj.save()
        url_VIDEO = 'https://api.negbuy.com' + product_obj.video.url
        return Response({
            'status': 'success',
            'message': 'video added',
            'data': url_VIDEO
        })
    except Exception as e:
        return Response({
            'status': 'error',
            'message': str(e)
        })


# Deepa's Size API
=======
@api_view(['POST'])
def api_RFQ(request):
    user_id = request.headers['User-id']
    required = request.data['required']
    target_price = request.data['target_price']
    try:
        user_obj= userDB.objects.get(user_id=user_id)
        rfq_db.objects.create(
            user= user_obj,
            required= required,
            target_price= target_price
        )
        data_dict= {
            'user': user_id,
            'required': required,
            'target_price': target_price
        }
        return Response({
            'status': True,
            'message': 'Success',
            'data' : data_dict
        })

    except Exception as e:
        return Response({
            'status': 'Error',
            'message': e,
        })

@api_view(['GET'])
def api_buyer_questions(request):
    user_id = request.headers['User-id']
    product_id = request.data['product_id']
    question = request.data['question']
    feedback = request.data['feedback']
    try:
        user_obj= userDB.objects.get(user_id=user_id)
        product_obj= product.objects.get(id=product_id)
        buyer_questions.objects.create(
            user= user_obj,
            product= product_obj, 
            question= question,
            # feedback= feedback
        )
        data_dict= {
            # 'user': user,
            # 'product': product,
            'question': question,
            'feedback':'' 
        }
        return Response({
            'status': True,
            'message': 'Success',
            'data' : data_dict
        })

    except Exception as e:
        return Response({
            'status': 'Error',
            'message': e,
        })


>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
@api_view(['GET'])
def size_api(request):
    user_id = request.headers['User-id']
    workbook = openpyxl.load_workbook('Dropdown_list.xlsx')
    ws = workbook['Sheet1']

    items = []
<<<<<<< HEAD
    a = []
    for row in ws.iter_rows(1, ):
        row_cells = []
=======
    a=[]
    for row in ws.iter_rows(1, ):
        row_cells=[]
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
        for cell in row:
            row_cells.append(cell.value)
        items.append(tuple(row_cells))

    for i in range(1, len(items)):
<<<<<<< HEAD
        s = items[i][1]
        a.append(s)

    return Response({
        'status': True,
        'message': 'Success',
        'data': a
    })


@api_view(['POST'])
def update_inventory(request):
    data = request.data
    inventory_obj = Inventory.objects.get(
        product__id=data['product_id'], color=data['color'])
    inventory_obj.quantity = data['quantity']
    inventory_obj.size = ','.join(data['size'])
    inventory_obj.save()

    return Response({
        'status': True,
        'message': 'Success',
    })


@api_view(['GET'])
def get_inventory(request):
    user_id = request.headers['User-id']
    user_products = product.objects.filter(
        user__user_id=user_id).values_list('id')

    return_list = []
    for item in user_products:
        print(item)
        data_dict = {}
        sum = 0
        inventory_obj = list(Inventory.objects.filter(
            product=item[0]).values('quantity'))
        for q in inventory_obj:
            sum = sum + q['quantity']

        product_obj = list(product.objects.filter(
            id=item[0]).values('name', 'price', 'id'))
        product_obj[0]['quantity'] = sum
        return_list.append(product_obj[0])

    return Response({
        'status': True,
        'message': 'Success',
        'data': return_list
    })

=======
        s=items[i][1]
        a.append(s)

    return Response({
         'status': True,
         'message': 'Success',
         'data' : a
     })
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5

@api_view(['GET'])
def db_rfq(request):
    user_id = request.headers['User-id']
<<<<<<< HEAD
    requirement = request.data['requirement']
    quantity = request.data['quantity']
    target_price = request.data['target_price']
    try:
        user_obj = userDB.objects.get(user_id=user_id)
        rfq.objects.create(
            user=user_obj,
            requirement=requirement,
            quantity=quantity,
            target_price=target_price
        )
        data_dict = {
            'user': user_id,
            'requirement': requirement,
            'quantity': quantity,
            'target_price': target_price
        }
        return Response({
            'status': True,
            'message': 'Success',
            'data': data_dict
=======
    all_rfq_list= list()
    try:
        user_obj= userDB.objects.get(user_id=user_id)
        all_rfq=rfq.objects.all()
        for each_rfq in all_rfq:
            obj= {
                'user': each_rfq.user_id,
                'requirement': each_rfq.requirement,
                'quantity' : each_rfq.quantity,
                'target_price': each_rfq.target_price
            }
            all_rfq_list.append(obj)
        return Response({
            'status': True,
            'message': 'Success',
            'data' : all_rfq_list
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
        })

    except Exception as e:
        return Response({
            'status': 'Error',
            'message': e,
        })

<<<<<<< HEAD

@api_view(['POST'])
def inventory_detail(request):
    # try:
    product_id = request.data['product_id']
    inventory_list = Inventory.objects.filter(product=product_id)
    product_obj = product.objects.get(id=product_id)

    if product_obj.varients:
        product_serialized = product_serializer(product_obj).data
        inventory_serialized = InventorySerializer(
            inventory_list, many=True).data
        product_serialized['varients'] = inventory_serialized
        if product_serialized['sku'] == None:
            product_serialized['sku'] = ''
        if product_serialized['size'] == None:
            product_serialized['size'] = ''

        return Response({
            'status': True,
            'message': 'Success',
            'data': product_serialized
        })
    else:
        product_serialized = product_serializer(product_obj).data
        inventory_obj = list(Inventory.objects.filter(
            product=product_id).values('size', 'quantity'))
        product_serialized['size'] = inventory_obj[0]['size']
        product_serialized['quantity'] = inventory_obj[0]['quantity']
        image_obj = productImages.objects.filter(product=product_id)
        product_images = ImageSerializer(image_obj, many=True).data
        product_serialized['varients'] = []
        product_serialized['images'] = product_images
        if product_serialized['sku'] == None:
            product_serialized['sku'] = ''
        if product_serialized['size'] == None:
            product_serialized['size'] = ''

        return Response({
            'status': True,
            'message': 'Success',
            'data': product_serialized
        })
    # except Exception as e:
    #     return Response({
    #         'status': False,
    #         'message': 'Error',
    #         'data': str(e)
    #     })
=======
>>>>>>> c571916195af49b2171c76e667617a10b583c0a5
